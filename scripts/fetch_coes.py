#!/usr/bin/env python3
"""Fetch REAL marginal cost from COES and overlay it onto the monthly history
(public/data/marginal_cost_history.json).

COES `ExportarMasivo` returns an .xlsx of half-hourly marginal cost per barra for
a date range — public, no auth, no token (discovered by reverse-engineering
costomarginal.js). We pull the last N complete months for the ≥138 kV backbone
barras (`defecto=S`), average TOTAL across all barras & half-hours per month,
convert soles→USD, and UPSERT those months onto the curated series (so the long
curated backdrop stays and recent months become real).

Endpoint: .../costosmarginales/ExportarMasivo?fechaInicio=DD/MM/YYYY&fechaFin=DD/MM/YYYY&defecto=S
Sheet COSTOMARGINAL, header row 5, data row 6+, columns: FECHA HORA | NODO EMD |
NOMBRE BARRA | ENERGÍA | CONGESTIÓN | TOTAL (S/./MWh).

Best-effort: months that fail are skipped; if nothing is fetched, exit 1 so the
pipeline keeps the curated series. Requires: requests, openpyxl.
"""

import calendar
import io
import json
import os
import sys
import time
import urllib.parse
from datetime import date

import openpyxl
import requests

from _meta import wrap

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, '..', 'public', 'data')
OUT = os.path.join(DATA, 'marginal_cost_history.json')

BASE = 'https://www.coes.org.pe/Portal/mercadomayorista/costosmarginales/ExportarMasivo'
PEN_PER_USD = 3.75  # soles per USD (approx 2024-25); COES publishes CMg in soles
MONTHS_BACK = int(os.environ.get('COES_MONTHS_BACK', '6'))
DEFECTO = 'S'  # backbone barras (>=138 kV); '' would pull all ~194 nodes
TIMEOUT = 90
RETRIES = 2


def fetch_chunk(d0: date, d1: date):
    """Return (sum_total_soles, count) of TOTAL over the [d0, d1] window."""
    params = {
        'fechaInicio': d0.strftime('%d/%m/%Y'),
        'fechaFin': d1.strftime('%d/%m/%Y'),
        'defecto': DEFECTO,
    }
    url = BASE + '?' + urllib.parse.urlencode(params, quote_via=urllib.parse.quote)
    last_err = None
    for attempt in range(RETRIES + 1):
        try:
            r = requests.get(url, timeout=TIMEOUT)
            r.raise_for_status()
            wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            total_col = None
            s = 0.0
            n = 0
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if i == 5:  # header
                    total_col = next(
                        (idx for idx, h in enumerate(row) if h and 'TOTAL' in str(h).upper()), 6
                    )
                    continue
                if i < 6 or not row:
                    continue
                v = row[total_col]
                if isinstance(v, (int, float)):
                    s += v
                    n += 1
            wb.close()
            return s, n
        except Exception as e:  # noqa: BLE001
            last_err = e
            time.sleep(1.5 * (attempt + 1))
    raise RuntimeError(f'chunk {d0}..{d1} failed: {last_err}')


def month_mean_usd(year: int, month: int):
    """Mean marginal cost (USD/MWh) for a month, via two <=16-day chunks."""
    last_day = calendar.monthrange(year, month)[1]
    chunks = [(date(year, month, 1), date(year, month, 15)),
              (date(year, month, 16), date(year, month, last_day))]
    tot_s, tot_n = 0.0, 0
    for d0, d1 in chunks:
        s, n = fetch_chunk(d0, d1)
        tot_s += s
        tot_n += n
        time.sleep(0.5)  # be polite
    if tot_n == 0:
        return None
    return round((tot_s / tot_n) / PEN_PER_USD, 1)


def target_months(n_back: int):
    """The n_back complete months ending with last month (skip the current partial one)."""
    today = date.today()
    y, m = today.year, today.month
    months = []
    for _ in range(n_back):
        m -= 1
        if m == 0:
            m = 12
            y -= 1
        months.append((y, m))
    return list(reversed(months))


def load_existing():
    if not os.path.exists(OUT):
        return []
    raw = json.load(open(OUT, encoding='utf-8'))
    return raw.get('data', raw) if isinstance(raw, dict) else raw


def main():
    fetched = {}
    for y, m in target_months(MONTHS_BACK):
        key = f'{y}-{m:02d}'
        try:
            usd = month_mean_usd(y, m)
            if usd is not None:
                fetched[key] = usd
                print(f'COES {key}: {usd} USD/MWh (real)')
        except Exception as e:  # noqa: BLE001
            print(f'skip {key}: {e}', file=sys.stderr)

    if not fetched:
        print('COES: no real data fetched — keeping curated series', file=sys.stderr)
        sys.exit(1)

    # Upsert real months onto the curated series.
    by_month = {r['month']: dict(r) for r in load_existing()}
    for key, usd in fetched.items():
        by_month[key] = {'month': key, 'cmg_usd_mwh': usd}
    rows = [by_month[k] for k in sorted(by_month)]

    envelope = wrap(
        rows,
        source=f'COES ExportarMasivo (últimos {len(fetched)} meses reales) + curado histórico',
        source_date=max(fetched),
        reliability='verified',
        note=f'Media de TOTAL sobre barras ≥138 kV, convertida a USD @ {PEN_PER_USD} S/./USD.',
    )
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(envelope, f, ensure_ascii=False, indent=2)
    print(f'\nupserted {len(fetched)} real months into marginal_cost_history.json')


if __name__ == '__main__':
    main()
